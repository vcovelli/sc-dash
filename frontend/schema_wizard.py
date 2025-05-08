import csv
import os
import requests
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Resolve paths relative to the current file
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_TEMPLATE = os.path.join(BASE_DIR, "../datasets/master/master_template.csv")
SCHEMA_CONFIG_DIR = os.path.join(BASE_DIR, "../user_schemas")
API_BASE_URL = os.getenv("API_BASE_URL", "http://localhost:8000")
TEMPLATE_OUTPUT_DIR = os.path.join(BASE_DIR, "../datasets/templates")

# Basic mapping of keywords to data types
def infer_postgres_type(column_name):
    name = column_name.lower()
    if any(kw in name for kw in ['date', 'time']):
        return "TIMESTAMP"
    elif any(kw in name for kw in ['id', 'quantity', 'amount', 'count']):
        return "INTEGER"
    elif any(kw in name for kw in ['price', 'cost', 'rate', 'value']):
        return "FLOAT"
    else:
        return "TEXT"

def read_master_template():
    with open(MASTER_TEMPLATE, mode='r') as file:
        reader = csv.reader(file)
        return next(reader)

def prompt_user_for_schema(headers):
    print("📊 Welcome to the Supply Chain Analytics Setup Wizard.\n")
    selected_columns = []

    for column in headers:
        include = input(f"Include column '{column}'? (y/n): ").strip().lower()
        if include == 'y':
            label = input(f"Custom label for '{column}' (or press Enter to keep as-is): ").strip() or column
            data_type = infer_postgres_type(column)
            selected_columns.append({
                "column_name": label,
                "data_type": data_type
            })

    return selected_columns

def save_schema_config(user_id, selected_columns):
    os.makedirs(SCHEMA_CONFIG_DIR, exist_ok=True)
    filepath = os.path.join(SCHEMA_CONFIG_DIR, f"{user_id}_schema.csv")
    with open(filepath, mode='w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=['column_name', 'data_type'])
        writer.writeheader()
        for col in selected_columns:
            writer.writerow(col)
    print(f"\nSchema saved to: {filepath}")

# Generate blank data template
def generate_template_xlsx(client_name, output_dir=TEMPLATE_OUTPUT_DIR, num_rows=10):
    schema_path = os.path.join(SCHEMA_CONFIG_DIR, f"{client_name}_schema.csv")
    output_path = os.path.join(output_dir, f"{client_name}_data_template.xlsx")

    if not os.path.exists(schema_path):
        raise FileNotFoundError(f"Schema not found: {schema_path}")

    os.makedirs(output_dir, exist_ok=True)

    # Read schema to get column names
    with open(schema_path, mode='r', newline='') as file:
        reader = csv.DictReader(file)
        columns = [row['column_name'] for row in reader]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{client_name}_data"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Write header
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # Add placeholder rows with UUIDs
    for row_idx in range(2, num_rows + 2):  # Start at row 2
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name == 'id':
                cell.value = str(uuid.uuid4())
            else:
                cell.value = ""

    # Autosize column widths
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = max(len(col_name), 36 if col_name == 'id' else 10)
        ws.column_dimensions[col_letter].width = max_length

    wb.save(output_path)
    print(f"Excel template with formulas saved at: {output_path}")

def trigger_table_creation(user_id):
    try:
        response = requests.post(
            f"{API_BASE_URL}/api/create-table/",
            json={"client_name": user_id}
        )
        if response.status_code == (200, 201):
            print(f"Table created successfully for `{user_id}`.")
        else:
            print(f"Table creation failed: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"Error contacting backend: {e}")

def main():
    user_id = input("Enter a unique name for this schema (e.g., company name): ").strip()
    headers = read_master_template()
    selected_columns = prompt_user_for_schema(headers)
    if selected_columns:
        if not any(col['column_name'] == 'id' for col in selected_columns):
            selected_columns.insert(0, {"column_name": "id", "data_type": "TEXT"})
        save_schema_config(user_id, selected_columns)
        trigger_table_creation(user_id)
        generate_template_xlsx(user_id)
    else:
        print("No columns selected. Schema not saved.")

if __name__ == "__main__":
    main()
