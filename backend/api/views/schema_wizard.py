import json, re, csv, os, random, uuid
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter
from minio import Minio
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from dotenv import load_dotenv
from api.views.table_creation import create_table_for_client
from api.models import UserSchema
from rest_framework_simplejwt.authentication import JWTAuthentication

load_dotenv()

SCHEMA_DIR = Path(os.environ.get("SCHEMA_DIR", "/opt/airflow/user_schemas"))
SCHEMA_DIR.mkdir(parents=True, exist_ok=True)

SCHEMA_FEATURES = {
    "orders": {
        "sheet": "MASTER_Orders",
        "columns": [
            "order_id", "order_date", "product_id", "product_name", "customer_id",
            "warehouse_id", "shipment_id", "quantity", "order_status", "expected_delivery_date"
        ]
    },
    "products": {
        "sheet": "Products",
        "columns": ["product_id", "product_name", "product_category", "supplier_id", "unit_price"]
    },
    "suppliers": {
        "sheet": "Suppliers",
        "columns": ["supplier_id", "supplier_name"]
    },
    "warehouses": {
        "sheet": "Warehouses",
        "columns": ["warehouse_id", "warehouse_location", "supplier_id"]
    },
    "customers": {
        "sheet": "Customers",
        "columns": ["customer_id", "customer_name"]
    },
    "shipments": {
        "sheet": "Shipments",
        "columns": ["shipment_id", "shipment_method", "shipment_status", "tracking_number"]
    }
}

REQUIRED_KEYS = ["order_id", "product_id", "order_date"]
INTERNAL_COLUMNS = ["version", "uuid", "ingested_at", "client_name"]


def style_header(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell.protection = Protection(locked=True)


def generate_sample_row(columns, client_name):
    row = []
    for col in columns:
        if col == "order_id":
            row.append(f"ORD-{random.randint(1000,9999)}")
        elif col == "product_id":
            row.append(f"PROD-{random.randint(100,999)}")
        elif col == "product_name":
            row.append(random.choice(["Widget A", "Widget B", "Widget C"]))
        elif col == "customer_id":
            row.append(f"CUST-{random.randint(100,999)}")
        elif col == "customer_name":
            row.append(random.choice(["Alice", "Bob", "Charlie", "Dana"]))
        elif col == "warehouse_id":
            row.append(f"WH-{random.randint(1,3)}")
        elif col == "warehouse_location":
            row.append(random.choice(["North DC", "South DC"]))
        elif col == "supplier_id":
            row.append(f"SUP-{random.randint(10,99)}")
        elif col == "supplier_name":
            row.append(random.choice(["SupplyCo", "MegaSupply", "FastParts"]))
        elif col == "shipment_id":
            row.append(f"SHIP-{random.randint(1000,9999)}")
        elif col == "shipment_method":
            row.append(random.choice(["Ground", "Air", "Freight"]))
        elif col == "shipment_status":
            row.append(random.choice(["Pending", "In Transit", "Delivered"]))
        elif col == "tracking_number":
            row.append(f"TRK{random.randint(100000,999999)}")
        elif col == "product_category":
            row.append(random.choice(["Electronics", "Apparel", "Furniture"]))
        elif col == "unit_price":
            row.append(round(random.uniform(10, 500), 2))
        elif col == "quantity":
            row.append(random.randint(1, 100))
        elif col == "order_date":
            row.append((datetime.today() - timedelta(days=random.randint(1, 30))).strftime("%Y-%m-%d"))
        elif col == "expected_delivery_date":
            row.append((datetime.today() + timedelta(days=random.randint(1, 14))).strftime("%Y-%m-%d"))
        elif col == "order_status":
            row.append(random.choice(["Pending", "Processing", "Delivered"]))
        elif col == "uuid":
            row.append(str(uuid.uuid4()))
        elif col == "version":
            row.append(1)
        elif col == "ingested_at":
            row.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        elif col == "client_name":
            row.append(client_name)
        else:
            row.append("Sample")
    return row

def write_sheet(wb, sheet_name, columns, client_name, num_rows=10):
    ws = wb.create_sheet(title=sheet_name)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)

    for row_num in range(2, 2 + num_rows):
        row_data = generate_sample_row(columns, client_name)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)

            # Set number formats
            if "date" in columns[col_idx-1]:
                cell.number_format = "yyyy-mm-dd"
            elif "price" in columns[col_idx-1]:
                cell.number_format = '"$"#,##0.00'

def generate_full_workbook(client_name, selected_features, include_sample_data=True):
    from openpyxl.formula.translate import Translator

    file_path = SCHEMA_DIR / f"{client_name}_full_template.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    included_sheets = {}
    for feature in selected_features:
        config = SCHEMA_FEATURES.get(feature)
        if config:
            included_sheets[config["sheet"]] = config["columns"]

    # Define logical order and remove backend-only fields
    logical_order = [
        "order_id", "order_date", "product_id", "product_name",
        "customer_id", "customer_name",
        "warehouse_id", "warehouse_location",
        "shipment_id", "shipment_method", "shipment_status", "tracking_number",
        "quantity", "unit_price", "product_category", "supplier_id", "supplier_name"
    ]
    master_fields = [col for col in logical_order if col in {col for cols in included_sheets.values() for col in cols} or col in REQUIRED_KEYS]

    # 1. MASTER_Orders Sheet with sample values
    ws_master = wb.create_sheet("MASTER_Orders")

    for col_idx, col_name in enumerate(master_fields, start=1):
        cell = ws_master.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)

    if include_sample_data:
        for row_num in range(2, 12):
            row_data = generate_sample_row(master_fields, client_name)
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_master.cell(row=row_num, column=col_idx, value=value)

            # Format cells
            if "date" in master_fields[col_idx - 1]:
                cell.number_format = "yyyy-mm-dd"
            elif "price" in master_fields[col_idx - 1]:
                cell.number_format = '"$"#,##0.00'

    # 2. Relational sheets using formulas
    for sheet_name, columns in included_sheets.items():
        if sheet_name == "MASTER_Orders":
            continue
        ws = wb.create_sheet(sheet_name)

        # Write headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            style_header(cell)

        # Relate each cell in this sheet to MASTER_Orders if the column exists there
        if include_sample_data:
            for row_num in range(2, 7):  # 5 rows
                for col_idx, col_name in enumerate(columns, start=1):
                    if col_name in master_fields:
                        ref_col_idx = master_fields.index(col_name) + 1
                        formula = f"'MASTER_Orders'!{get_column_letter(ref_col_idx)}{row_num}"
                        ws.cell(row=row_num, column=col_idx, value=f"={formula}")
                    else:
                        ws.cell(row=row_num, column=col_idx, value="Sample")

    wb.save(file_path)

    # Upload to MinIO
    minio_client = Minio(
        os.getenv("MINIO_ENDPOINT").replace("http://", "").replace("https://", ""),
        access_key=os.getenv("MINIO_ROOT_USER"),
        secret_key=os.getenv("MINIO_ROOT_PASSWORD"),
        secure=False,
    )

    bucket_name = "templates"
    object_name = file_path.name

    if not minio_client.bucket_exists(bucket_name):
        minio_client.make_bucket(bucket_name)

    minio_client.fput_object(bucket_name, object_name, str(file_path))

    url = minio_client.presigned_get_object(bucket_name, object_name, expires=timedelta(minutes=30))

    public_minio_url = os.getenv("PUBLIC_MINIO_URL")
    if public_minio_url:
        url = url.replace("http://minio:9000", public_minio_url)

    return {
        "download_url": url,
        "file_path": str(file_path)  # actual local .xlsx path
    }

@csrf_exempt
def generate_schema(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)

    try:
        data = json.loads(request.body)
        raw_client_name = data.get("client_name") or data.get("business_name")
        selected_features = data.get("features")
        include_sample_data = data.get("include_sample_data", True)

        if not raw_client_name or not selected_features:
            return JsonResponse({"error": "Missing client_name or features"}, status=400)

        client_name = re.sub(r'[^a-z0-9]', '', raw_client_name.lower())
        allowed_keys = set(SCHEMA_FEATURES.keys())
        selected_features = [f for f in selected_features if f in allowed_keys]

        if not selected_features:
            return JsonResponse({"error": "No valid features selected"}, status=400)

        # Step 1: Write schema CSV
        schema_path = SCHEMA_DIR / f"{client_name}_schema.csv"
        with open(schema_path, mode="w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["column_name", "data_type"])

            all_columns = set()
            for feature in selected_features:
                config = SCHEMA_FEATURES.get(feature)
                if config:
                    all_columns.update(config["columns"])

            all_columns.update(REQUIRED_KEYS)
            all_columns.update(INTERNAL_COLUMNS)

            for col in sorted(all_columns):
                if "date" in col:
                    dtype = "DATE"
                elif "id" in col or col == "uuid":
                    dtype = "VARCHAR(64)"
                elif col == "quantity":
                    dtype = "INTEGER"
                elif col == "unit_price":
                    dtype = "NUMERIC(10, 2)"
                else:
                    dtype = "TEXT"
                writer.writerow([col, dtype])

        # Step 2: Collect expected headers
        expected_headers_set = set()
        for feature in selected_features:
            config = SCHEMA_FEATURES.get(feature)
            if config:
                expected_headers_set.update(config["columns"])

        expected_headers_set.update(REQUIRED_KEYS)
        expected_headers = sorted(list(expected_headers_set))

        # Step 3: Auth & DB update
        user = None
        try:
            authenticator = JWTAuthentication()
            user_auth_tuple = authenticator.authenticate(request)
            if user_auth_tuple:
                user, _ = user_auth_tuple
                UserSchema.objects.update_or_create(
                    user=user,
                    defaults={"expected_headers": expected_headers}
                )
        except Exception as auth_err:
            print(f"[schema_wizard] ⚠️ User authentication failed: {auth_err}")

        # Step 4: Generate workbook (.xlsx)##############################################################################################################
        create_table_for_client(client_name)#############################################################################################################
        workbook = generate_full_workbook(client_name, selected_features, include_sampl##################################################################
        download_url = workbook["download_url"]##########################################################################################################

        # Step 5: Download from MinIO for Grist upload###################################################################################################
        import requests##################################################################################################################################
        local_temp_path = f"/tmp/{client_name}_upload.xlsx" # This all has to change to##################################################################
        r = requests.get(download_url) ##################################################################################################################
        with open(local_temp_path, "wb") as f: ##########################################################################################################
            f.write(r.content)###########################################################################################################################

        # Step 6: Generate .grist file from schema + sample data#########################################################################################
        schema, sample_data = generate_grist_schema_and_data(selected_features, client_##################################################################
        grist_path = f"/tmp/{client_name}.grist"#########################################################################################################
        create_grist_file(grist_path, schema, sample_data)###############################################################################################
        grist_doc = upload_grist_file(grist_path)########################################################################################################

        # Step 7: Save Grist view to DB if authenticated#################################################################################################
        if user:##################################################################
            UserSchema.objects.update_or_create(#########################################################################################################
                user=user,###############################################################################################################################
                defaults={###############################################################################################################################
                    "expected_headers": expected_headers,################################################################################################
                    "grist_doc_id": grist_doc["doc_id"],#################################################################################################
                    "grist_doc_url": grist_doc["doc_url"],###############################################################################################
                    "grist_view_url": grist_doc["view_url"]##############################################################################################
                }####################################################################################################################################
            )

        return JsonResponse({
            "success": True,
            "message": f"Workbook and Grist doc generated for {client_name}",
            "download_url": download_url,
            "grist_doc_url": grist_doc["doc_url"],
            "grist_view_url": grist_doc["view_url"]
        })

    except Exception as e:
        print(f"[schema_wizard] ❌ Error: {e}")
        return JsonResponse({"error": str(e)}, status=500)