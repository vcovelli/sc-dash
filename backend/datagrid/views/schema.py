import json
import re
import csv
import os
import random
import uuid
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter
from helpers.table_utils import create_table_for_org
from helpers.minio_client import get_minio_client, ensure_bucket_exists

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from django.db.models import Q
from dotenv import load_dotenv

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import action

from datagrid.models import UserTableSchema, Column, SchemaHistory, SchemaPermission  # Use your app/model import here!
from datagrid.serializers import (
    UserTableSchemaSerializer, UserTableSchemaCreateUpdateSerializer, 
    ColumnSerializer, SchemaHistorySerializer, SchemaPermissionSerializer
)  # Ditto, update path if needed

from accounts.models import UserActivity  # If you track user actions
from accounts.permissions import (
    IsReadOnlyOrAbove, CanViewAnalytics, CanCreateSchemas, 
    CanShareSchemas, CanAccessSharedSchemas, CanAccessSchema, CanEditSchema
)
from accounts.mixins import CombinedOrgMixin
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.contrib.auth import get_user_model
from django.db import models

User = get_user_model()

# ====== SCHEMA CONSTANTS AND HELPERS =======

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent.parent.parent
SCHEMA_DIR = Path(os.environ.get("SCHEMA_DIR", BASE_DIR / "user_schemas"))

def ensure_schema_dir():
    SCHEMA_DIR.mkdir(parents=True, exist_ok=True)

SCHEMA_FEATURES = {
    "orders": {
        "sheet": "MASTER_Orders",
        "columns": [
            "order_id", "order_date", "product_id", "product_name", "customer_id",
            "warehouse_id", "shipment_id", "quantity", "order_status", "expected_delivery_date"
        ]
    },
    "products": {
        "sheet": "Products",
        "columns": ["product_id", "product_name", "product_category", "supplier_id", "unit_price"]
    },
    "suppliers": {
        "sheet": "Suppliers",
        "columns": ["supplier_id", "supplier_name"]
    },
    "warehouses": {
        "sheet": "Warehouses",
        "columns": ["warehouse_id", "warehouse_location", "supplier_id"]
    },
    "customers": {
        "sheet": "Customers",
        "columns": ["customer_id", "customer_name"]
    },
    "shipments": {
        "sheet": "Shipments",
        "columns": ["shipment_id", "shipment_method", "shipment_status", "tracking_number"]
    }
}

REQUIRED_KEYS = ["order_id", "product_id", "order_date"]
INTERNAL_COLUMNS = ["version", "uuid", "ingested_at", "client_name"]

def style_header(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell.protection = Protection(locked=True)

def generate_sample_row(columns, client_name):
    row = []
    for col in columns:
        if col == "order_id":
            row.append(f"ORD-{random.randint(1000,9999)}")
        elif col == "product_id":
            row.append(f"PROD-{random.randint(100,999)}")
        elif col == "product_name":
            row.append(random.choice(["Widget A", "Widget B", "Widget C"]))
        elif col == "customer_id":
            row.append(f"CUST-{random.randint(100,999)}")
        elif col == "customer_name":
            row.append(random.choice(["Alice", "Bob", "Charlie", "Dana"]))
        elif col == "warehouse_id":
            row.append(f"WH-{random.randint(1,3)}")
        elif col == "warehouse_location":
            row.append(random.choice(["North DC", "South DC"]))
        elif col == "supplier_id":
            row.append(f"SUP-{random.randint(10,99)}")
        elif col == "supplier_name":
            row.append(random.choice(["SupplyCo", "MegaSupply", "FastParts"]))
        elif col == "shipment_id":
            row.append(f"SHIP-{random.randint(1000,9999)}")
        elif col == "shipment_method":
            row.append(random.choice(["Ground", "Air", "Freight"]))
        elif col == "shipment_status":
            row.append(random.choice(["Pending", "In Transit", "Delivered"]))
        elif col == "tracking_number":
            row.append(f"TRK{random.randint(100000,999999)}")
        elif col == "product_category":
            row.append(random.choice(["Electronics", "Apparel", "Furniture"]))
        elif col == "unit_price":
            row.append(round(random.uniform(10, 500), 2))
        elif col == "quantity":
            row.append(random.randint(1, 100))
        elif col == "order_date":
            row.append((datetime.today() - timedelta(days=random.randint(1, 30))).strftime("%Y-%m-%d"))
        elif col == "expected_delivery_date":
            row.append((datetime.today() + timedelta(days=random.randint(1, 14))).strftime("%Y-%m-%d"))
        elif col == "order_status":
            row.append(random.choice(["Pending", "Processing", "Delivered"]))
        elif col == "uuid":
            row.append(str(uuid.uuid4()))
        elif col == "version":
            row.append(1)
        elif col == "ingested_at":
            row.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        elif col == "client_name":
            row.append(client_name)
        else:
            row.append("Sample")
    return row

def write_sheet(wb, sheet_name, columns, client_name, num_rows=10):
    ws = wb.create_sheet(title=sheet_name)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)
    for row_num in range(2, 2 + num_rows):
        row_data = generate_sample_row(columns, client_name)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            if "date" in columns[col_idx-1]:
                cell.number_format = "yyyy-mm-dd"
            elif "price" in columns[col_idx-1]:
                cell.number_format = '"$"#,##0.00'

def generate_full_workbook(client_name, selected_features, include_sample_data=True):
    ensure_schema_dir()
    file_path = SCHEMA_DIR / f"{client_name}_full_template.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    included_sheets = {}
    for feature in selected_features:
        config = SCHEMA_FEATURES.get(feature)
        if config:
            included_sheets[config["sheet"]] = config["columns"]

    logical_order = [
        "order_id", "order_date", "product_id", "product_name",
        "customer_id", "customer_name",
        "warehouse_id", "warehouse_location",
        "shipment_id", "shipment_method", "shipment_status", "tracking_number",
        "quantity", "unit_price", "product_category", "supplier_id", "supplier_name"
    ]
    master_fields = [col for col in logical_order if col in {col for cols in included_sheets.values() for col in cols} or col in REQUIRED_KEYS]

    # MASTER_Orders Sheet with sample values
    ws_master = wb.create_sheet("MASTER_Orders")
    for col_idx, col_name in enumerate(master_fields, start=1):
        cell = ws_master.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)
    if include_sample_data:
        for row_num in range(2, 12):
            row_data = generate_sample_row(master_fields, client_name)
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_master.cell(row=row_num, column=col_idx, value=value)
                if "date" in master_fields[col_idx - 1]:
                    cell.number_format = "yyyy-mm-dd"
                elif "price" in master_fields[col_idx - 1]:
                    cell.number_format = '"$"#,##0.00'

    # Relational sheets using formulas
    for sheet_name, columns in included_sheets.items():
        if sheet_name == "MASTER_Orders":
            continue
        ws = wb.create_sheet(sheet_name)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            style_header(cell)
        if include_sample_data:
            for row_num in range(2, 7):
                for col_idx, col_name in enumerate(columns, start=1):
                    if col_name in master_fields:
                        ref_col_idx = master_fields.index(col_name) + 1
                        formula = f"'MASTER_Orders'!{get_column_letter(ref_col_idx)}{row_num}"
                        ws.cell(row=row_num, column=col_idx, value=f"={formula}")
                    else:
                        ws.cell(row=row_num, column=col_idx, value="Sample")

    wb.save(file_path)

    # --- MinIO connection using centralized configuration ---
    minio_client = get_minio_client()
    bucket_name = "templates"
    object_name = file_path.name

    # --- Ensure bucket exists ---
    ensure_bucket_exists(minio_client, bucket_name)

    # --- Upload file ---
    minio_client.fput_object(bucket_name, object_name, str(file_path))

    # --- Presigned URL using public host, no netloc swap needed ---
    presigned_url = minio_client.presigned_get_object(
        bucket_name, object_name, expires=timedelta(minutes=30)
    )

    return {
        "download_url": presigned_url,
        "file_path": str(file_path)
    }

@csrf_exempt
def generate_schema(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)

    try:
        # --- 1. Parse Request Data ---
        data = json.loads(request.body)
        raw_client_name = data.get("client_name") or data.get("business_name") or data.get("client_id")
        selected_features = data.get("features")
        include_sample_data = data.get("include_sample_data", True)

        if not raw_client_name or not selected_features:
            return JsonResponse({"error": "Missing client_name/client_id or features"}, status=400)

        client_name = re.sub(r'[^a-z0-9]', '', raw_client_name.lower())
        allowed_keys = set(SCHEMA_FEATURES.keys())
        selected_features = [f for f in selected_features if f in allowed_keys]

        if not selected_features:
            return JsonResponse({"error": "No valid features selected"}, status=400)

        # --- 2. Compute All Columns (Union) ---
        all_columns = set()
        for feature in selected_features:
            config = SCHEMA_FEATURES.get(feature)
            if config:
                all_columns.update(config["columns"])
        all_columns.update(REQUIRED_KEYS)
        all_columns.update(INTERNAL_COLUMNS)
        sorted_columns = sorted(all_columns)

        # --- 3. Write schema CSV (optional, just for export/debug) ---
        ensure_schema_dir()
        schema_path = SCHEMA_DIR / f"{client_name}_schema.csv"
        with open(schema_path, mode="w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["column_name", "data_type"])
            for col in sorted_columns:
                if "date" in col:
                    dtype = "DATE"
                elif "id" in col or col == "uuid":
                    dtype = "VARCHAR(64)"
                elif col == "quantity":
                    dtype = "INTEGER"
                elif col == "unit_price":
                    dtype = "NUMERIC(10, 2)"
                else:
                    dtype = "TEXT"
                writer.writerow([col, dtype])

        # --- 4. Authenticate User ---
        user = None
        try:
            authenticator = JWTAuthentication()
            user_auth_tuple = authenticator.authenticate(request)
            if user_auth_tuple:
                user, _ = user_auth_tuple
        except Exception as auth_err:
            print(f"[schema_wizard] ⚠️ User authentication failed: {auth_err}")

        # --- 5. Save User Schema (if authenticated) ---
        # Ensure user has an organization
        if user and not user.org:
            from accounts.models import Organization
            org_name = f"{user.first_name or user.username}'s Organization"
            org, _ = Organization.objects.get_or_create(
                name=org_name,
                defaults={"slug": f"{user.username}-org"}
            )
            user.org = org
            user.save()
             

        if user and user.org:
            # Log onboarding complete
            UserActivity.objects.create(
                user=user,
                verb="completed onboarding",
                target=client_name,
                meta={
                    "features": selected_features,
                    "include_sample_data": include_sample_data,
                }
            )
            for feature in selected_features:
                config = SCHEMA_FEATURES.get(feature)
                if config:
                    table_name = feature
                    columns = list(config["columns"])
                    # Add any global or table-specific columns as needed
                    columns += [c for c in (REQUIRED_KEYS + INTERNAL_COLUMNS) if c not in columns]
                    columns = sorted(set(columns))
                    UserTableSchema.objects.update_or_create(
                        user=user,
                        org=user.org,
                        table_name=table_name,
                        defaults={"columns": columns}
                    )

        # --- 6. Generate the Excel Workbook, Upload to Minio ---
        # If you have your own client table creation, call it here:
        create_table_for_org(client_name)
        workbook = generate_full_workbook(client_name, selected_features, include_sample_data)
        download_url = workbook["download_url"]

        # Log template download if user is authenticated
        if user and download_url:
            UserActivity.objects.create(
                user=user,
                verb="downloaded template",
                target=os.path.basename(download_url),
                meta={"download_url": download_url, "client_name": client_name}
            )

        # --- 7. Respond with Success/Download URL ---
        return JsonResponse({
            "success": True,
            "message": f"Workbook generated for {client_name}",
            "download_url": download_url
        })

    except Exception as e:
        print(f"[schema_wizard] ❌ Error: {e}")
        return JsonResponse({"error": str(e)}, status=500)

# ==== REST API: UserTableSchema CRUD (multi-table support) ====

def infer_type(col_name):
    if col_name.endswith("_id"):
        return "reference"
    if col_name.endswith("_ids"):
        return "reference-multi"
    if "date" in col_name:
        return "date"
    if "status" in col_name:
        return "choice"
    if "quantity" in col_name or "amount" in col_name or "number" in col_name:
        return "number"
    return "text"

def normalize_columns(columns):
    if not columns:
        return []
    if isinstance(columns[0], dict):
        for col in columns:
            if "type" not in col:
                col["type"] = infer_type(col.get("accessorKey") or col.get("header", ""))
        return columns
    out = []
    for col in columns:
        out.append({
            "accessorKey": col,
            "header": col.replace("_", " ").title(),
            "type": infer_type(col)
        })
    return out

class UserTableSchemasView(CombinedOrgMixin, APIView):
    """Handle CRUD for user table schemas with organization-wide sharing support."""
    permission_classes = [CanCreateSchemas]

    def get(self, request):
        """
        Get all accessible table schemas for this user.
        Includes personal schemas and organization-wide shared schemas.
        """
        # Base queryset - schemas in user's organization
        base_query = UserTableSchema.objects.filter(org=request.user.org)
        
        # Get personal schemas + shared schemas user can access
        accessible_schemas = base_query.filter(
            Q(user=request.user) |  # Own schemas
            Q(sharing_level='organization', is_shared=True)  # Shared schemas
        ).distinct()
        
        # Check if user has permission to access shared schemas
        if request.user.role not in ['admin', 'owner', 'ceo', 'national_manager', 
                                   'regional_manager', 'local_manager', 'employee', 'client']:
            # Read-only users can only see their own schemas
            accessible_schemas = base_query.filter(user=request.user)
        
        # Add query parameter to filter by sharing level
        sharing_filter = request.query_params.get('sharing_level')
        if sharing_filter:
            if sharing_filter == 'personal':
                accessible_schemas = accessible_schemas.filter(user=request.user, sharing_level='personal')
            elif sharing_filter == 'shared':
                accessible_schemas = accessible_schemas.filter(sharing_level='organization', is_shared=True)
        
        # Normalize every schema's columns
        for schema in accessible_schemas:
            schema.columns = normalize_columns(schema.columns)
        
        serializer = UserTableSchemaSerializer(accessible_schemas, many=True, context={'request': request})
        return Response(serializer.data)

    def post(self, request):
        """Create a new table schema for this user."""
        table_name = request.data.get("table_name")
        columns = request.data.get("columns")
        sharing_level = request.data.get("sharing_level", "personal")

        if not table_name or not columns:
            return Response({"error": "Missing table_name or columns."}, status=400)
        
        # Check if schema already exists for this user
        if UserTableSchema.objects.filter(
            user=request.user, 
            org=request.user.org, 
            table_name=table_name
        ).exists():
            return Response({"error": "Schema already exists. Use PATCH to update."}, status=409)

        # Validate sharing level permissions
        if sharing_level == 'organization':
            SCHEMA_SHARE_ROLES = [
                'admin', 'owner', 'ceo', 'national_manager', 
                'regional_manager', 'local_manager'
            ]
            if request.user.role not in SCHEMA_SHARE_ROLES:
                return Response({
                    "error": "You don't have permission to create organization-wide schemas. "
                            "Create a personal schema instead."
                }, status=403)

        columns = normalize_columns(columns)
        serializer = UserTableSchemaSerializer(data={
            "table_name": table_name,
            "columns": columns,
            "sharing_level": sharing_level,
        }, context={'request': request})
        
        if serializer.is_valid():
            schema = serializer.save(user=request.user, org=request.user.org)
            
            # If creating as organization-wide, set sharing fields
            if sharing_level == 'organization':
                schema.share_organization_wide(request.user)
            
            return Response(UserTableSchemaSerializer(schema, context={'request': request}).data, status=201)
        return Response(serializer.errors, status=400)


class UserTableSchemaDetailView(CombinedOrgMixin, APIView):
    """Handle GET, PATCH, DELETE for a single table schema with sharing support."""
    permission_classes = [CanCreateSchemas]

    def get_object(self, request, table_name):
        """Get schema that user can access (own or shared)."""
        # Try to get user's own schema first
        try:
            return UserTableSchema.objects.get(
                user=request.user, 
                org=request.user.org, 
                table_name=table_name
            )
        except UserTableSchema.DoesNotExist:
            # If not found, try to get shared schema
            try:
                shared_schema = UserTableSchema.objects.get(
                    org=request.user.org,
                    table_name=table_name,
                    sharing_level='organization',
                    is_shared=True
                )
                # Check if user has permission to access shared schema
                if request.user.role in ['admin', 'owner', 'ceo', 'national_manager', 
                                       'regional_manager', 'local_manager', 'employee', 'client']:
                    return shared_schema
            except UserTableSchema.DoesNotExist:
                pass
            
            # Schema not found or no permission
            return None

    def get(self, request, table_name):
        schema = self.get_object(request, table_name)
        if not schema:
            return Response({"error": "Schema not found or access denied."}, status=404)
        
        schema.columns = normalize_columns(schema.columns)
        serializer = UserTableSchemaSerializer(schema, context={'request': request})
        return Response(serializer.data)

    def patch(self, request, table_name):
        schema = self.get_object(request, table_name)
        if not schema:
            return Response({"error": "Schema not found or access denied."}, status=404)
        
        # Check edit permissions
        if not schema.can_user_edit(request.user):
            return Response({
                "error": "You don't have permission to edit this schema."
            }, status=403)
        
        columns = request.data.get("columns")
        if columns is not None:
            columns = normalize_columns(columns)
            request.data["columns"] = columns
        
        serializer = UserTableSchemaSerializer(
            schema, 
            data=request.data, 
            partial=True, 
            context={'request': request}
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def delete(self, request, table_name):
        schema = self.get_object(request, table_name)
        if not schema:
            return Response({"error": "Schema not found or access denied."}, status=404)
        
        # Only schema owner can delete
        if schema.user != request.user:
            return Response({
                "error": "Only the schema owner can delete it."
            }, status=403)
        
        schema.delete()
        return Response({"success": True})

# ====== ENHANCED SCHEMA VIEWS ======

class EnhancedSchemasView(CombinedOrgMixin, APIView):
    """Enhanced schema management with versioning, sharing, and validation"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsReadOnlyOrAbove]

    def get(self, request):
        """Get schemas with enhanced filtering and sharing logic"""
        try:
            org = self.get_user_org(request.user)
            
            # Base queryset - user's schemas + shared schemas
            user_schemas = UserTableSchema.objects.filter(
                user=request.user, 
                org=org,
                is_active=True
            )
            
            # Add organization shared schemas
            org_shared = UserTableSchema.objects.filter(
                org=org,
                sharing_level='org',
                is_active=True
            ).exclude(user=request.user)
            
            # Add public schemas
            public_schemas = UserTableSchema.objects.filter(
                sharing_level='public',
                is_active=True
            ).exclude(user=request.user)
            
            # Combine querysets
            all_schemas = user_schemas.union(org_shared, public_schemas)
            
            # Apply additional filters
            table_name = request.GET.get('table_name')
            if table_name:
                all_schemas = all_schemas.filter(table_name__icontains=table_name)
            
            show_invalid = request.GET.get('show_invalid', 'false').lower() == 'true'
            if not show_invalid:
                all_schemas = all_schemas.filter(is_valid=True)
            
            # Order by sharing level (own first), then by name
            all_schemas = all_schemas.order_by('sharing_level', 'table_name')
            
            serializer = UserTableSchemaSerializer(all_schemas, many=True, context={'request': request})
            return Response(serializer.data)
            
        except Exception as e:
            return Response(
                {"error": f"Failed to retrieve schemas: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request):
        """Create new schema with enhanced features"""
        try:
            org = self.get_user_org(request.user)
            
            # Add user and org to the data
            data = request.data.copy()
            
            serializer = UserTableSchemaCreateUpdateSerializer(
                data=data, 
                context={'request': request}
            )
            
            if serializer.is_valid():
                schema = serializer.save(user=request.user, org=org)
                
                # Return full schema data
                response_serializer = UserTableSchemaSerializer(
                    schema, 
                    context={'request': request}
                )
                return Response(response_serializer.data, status=status.HTTP_201_CREATED)
            
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            return Response(
                {"error": f"Failed to create schema: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class EnhancedSchemaDetailView(CombinedOrgMixin, APIView):
    """Enhanced individual schema management"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsReadOnlyOrAbove]

    def get_schema(self, request, schema_id):
        """Get schema with permission checking"""
        try:
            schema = UserTableSchema.objects.get(id=schema_id)
            
            # Check if user can access this schema
            if schema.user == request.user:
                return schema
            elif schema.sharing_level == 'org' and schema.org == self.get_user_org(request.user):
                return schema
            elif schema.sharing_level == 'public':
                return schema
            elif schema.permissions.filter(user=request.user).exists():
                return schema
            else:
                return None
        except UserTableSchema.DoesNotExist:
            return None

    def get(self, request, schema_id):
        """Get detailed schema information"""
        schema = self.get_schema(request, schema_id)
        if not schema:
            return Response(
                {"error": "Schema not found or access denied"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Include history in detail view
        serializer = UserTableSchemaSerializer(schema, context={'request': request})
        return Response(serializer.data)

    def put(self, request, schema_id):
        """Update schema with change tracking"""
        schema = self.get_schema(request, schema_id)
        if not schema:
            return Response(
                {"error": "Schema not found or access denied"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check edit permissions
        can_edit = (
            schema.user == request.user or
            schema.permissions.filter(
                user=request.user, 
                permission__in=['edit', 'admin']
            ).exists()
        )
        
        if not can_edit:
            return Response(
                {"error": "Permission denied - cannot edit this schema"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = UserTableSchemaCreateUpdateSerializer(
            schema, 
            data=request.data, 
            partial=True,
            context={'request': request}
        )
        
        if serializer.is_valid():
            updated_schema = serializer.save()
            response_serializer = UserTableSchemaSerializer(
                updated_schema, 
                context={'request': request}
            )
            return Response(response_serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, schema_id):
        """Delete schema (soft delete by marking inactive)"""
        schema = self.get_schema(request, schema_id)
        if not schema:
            return Response(
                {"error": "Schema not found or access denied"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Only owner or admin can delete
        can_delete = (
            schema.user == request.user or
            schema.permissions.filter(
                user=request.user, 
                permission='admin'
            ).exists()
        )
        
        if not can_delete:
            return Response(
                {"error": "Permission denied - cannot delete this schema"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Soft delete
        schema.is_active = False
        schema.save()
        
        # Create history entry
        SchemaHistory.objects.create(
            schema=schema,
            action='deleted',
            user=request.user,
            description=f"Deleted schema '{schema.table_name}'"
        )
        
        return Response(status=status.HTTP_204_NO_CONTENT)


class SchemaVersionView(CombinedOrgMixin, APIView):
    """Manage schema versions"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsReadOnlyOrAbove]

    def post(self, request, schema_id):
        """Create new version of schema"""
        try:
            schema = UserTableSchema.objects.get(id=schema_id)
            
            # Check permissions
            if schema.user != request.user:
                return Response(
                    {"error": "Only schema owner can create versions"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            new_version = schema.create_new_version()
            
            # Create history entry
            SchemaHistory.objects.create(
                schema=new_version,
                action='version_created',
                user=request.user,
                description=f"Created version {new_version.version} from {schema.version}"
            )
            
            serializer = UserTableSchemaSerializer(new_version, context={'request': request})
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except UserTableSchema.DoesNotExist:
            return Response(
                {"error": "Schema not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to create version: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def get(self, request, schema_id):
        """Get all versions of a schema"""
        try:
            schema = UserTableSchema.objects.get(id=schema_id)
            
            # Get all versions
            versions = UserTableSchema.objects.filter(
                user=schema.user,
                org=schema.org,
                table_name=schema.table_name
            ).order_by('-version')
            
            serializer = UserTableSchemaSerializer(versions, many=True, context={'request': request})
            return Response(serializer.data)
            
        except UserTableSchema.DoesNotExist:
            return Response(
                {"error": "Schema not found"},
                status=status.HTTP_404_NOT_FOUND
            )


class SchemaColumnsView(CombinedOrgMixin, APIView):
    """Manage schema columns with drag-drop ordering"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsReadOnlyOrAbove]

    def get(self, request, schema_id):
        """Get columns for a schema"""
        try:
            schema = UserTableSchema.objects.get(id=schema_id)
            columns = schema.schema_columns.all().order_by('order')
            serializer = ColumnSerializer(columns, many=True)
            return Response(serializer.data)
        except UserTableSchema.DoesNotExist:
            return Response(
                {"error": "Schema not found"},
                status=status.HTTP_404_NOT_FOUND
            )

    def post(self, request, schema_id):
        """Add new column to schema"""
        try:
            schema = UserTableSchema.objects.get(id=schema_id)
            
            # Check edit permissions
            if schema.user != request.user:
                return Response(
                    {"error": "Permission denied"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Set order if not provided
            if 'order' not in request.data:
                max_order = schema.schema_columns.aggregate(
                    max_order=models.Max('order')
                )['max_order'] or 0
                request.data['order'] = max_order + 1
            
            serializer = ColumnSerializer(data=request.data)
            if serializer.is_valid():
                column = serializer.save(schema=schema)
                
                # Re-validate schema
                schema.validate_schema()
                
                # Create history entry
                SchemaHistory.objects.create(
                    schema=schema,
                    action='column_added',
                    user=request.user,
                    description=f"Added column '{column.name}'"
                )
                
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
        except UserTableSchema.DoesNotExist:
            return Response(
                {"error": "Schema not found"},
                status=status.HTTP_404_NOT_FOUND
            )

    def put(self, request, schema_id):
        """Reorder columns (drag-drop support)"""
        try:
            schema = UserTableSchema.objects.get(id=schema_id)
            
            # Check edit permissions
            if schema.user != request.user:
                return Response(
                    {"error": "Permission denied"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Expect array of column IDs in new order
            column_ids = request.data.get('column_order', [])
            
            for i, column_id in enumerate(column_ids):
                Column.objects.filter(
                    id=column_id, 
                    schema=schema
                ).update(order=i)
            
            # Create history entry
            SchemaHistory.objects.create(
                schema=schema,
                action='column_reordered',
                user=request.user,
                description="Reordered columns"
            )
            
            # Return updated columns
            columns = schema.schema_columns.all().order_by('order')
            serializer = ColumnSerializer(columns, many=True)
            return Response(serializer.data)
            
        except UserTableSchema.DoesNotExist:
            return Response(
                {"error": "Schema not found"},
                status=status.HTTP_404_NOT_FOUND
            )


class SchemaColumnDetailView(CombinedOrgMixin, APIView):
    """Individual column management"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsReadOnlyOrAbove]

    def get(self, request, schema_id, column_id):
        """Get column details"""
        try:
            column = Column.objects.get(id=column_id, schema__id=schema_id)
            serializer = ColumnSerializer(column)
            return Response(serializer.data)
        except Column.DoesNotExist:
            return Response(
                {"error": "Column not found"},
                status=status.HTTP_404_NOT_FOUND
            )

    def put(self, request, schema_id, column_id):
        """Update column"""
        try:
            column = Column.objects.get(id=column_id, schema__id=schema_id)
            
            # Check edit permissions
            if column.schema.user != request.user:
                return Response(
                    {"error": "Permission denied"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            old_data = ColumnSerializer(column).data
            
            serializer = ColumnSerializer(column, data=request.data, partial=True)
            if serializer.is_valid():
                updated_column = serializer.save()
                
                # Re-validate schema
                column.schema.validate_schema()
                
                # Create history entry
                SchemaHistory.objects.create(
                    schema=column.schema,
                    action='column_updated',
                    user=request.user,
                    field_changed=column.name,
                    old_value=old_data,
                    new_value=serializer.data,
                    description=f"Updated column '{column.name}'"
                )
                
                return Response(serializer.data)
            
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
        except Column.DoesNotExist:
            return Response(
                {"error": "Column not found"},
                status=status.HTTP_404_NOT_FOUND
            )

    def delete(self, request, schema_id, column_id):
        """Delete column"""
        try:
            column = Column.objects.get(id=column_id, schema__id=schema_id)
            
            # Check edit permissions
            if column.schema.user != request.user:
                return Response(
                    {"error": "Permission denied"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            column_name = column.name
            schema = column.schema
            column.delete()
            
            # Re-validate schema
            schema.validate_schema()
            
            # Create history entry
            SchemaHistory.objects.create(
                schema=schema,
                action='column_deleted',
                user=request.user,
                description=f"Deleted column '{column_name}'"
            )
            
            return Response(status=status.HTTP_204_NO_CONTENT)
            
        except Column.DoesNotExist:
            return Response(
                {"error": "Column not found"},
                status=status.HTTP_404_NOT_FOUND
            )


class SchemaSharingView(CombinedOrgMixin, APIView):
    """Manage schema sharing and permissions"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsReadOnlyOrAbove]

    def post(self, request, schema_id):
        """Share schema or grant permissions"""
        try:
            schema = UserTableSchema.objects.get(id=schema_id)
            
            # Only owner or admin can manage sharing
            can_share = (
                schema.user == request.user or
                schema.permissions.filter(
                    user=request.user, 
                    permission='admin'
                ).exists()
            )
            
            if not can_share:
                return Response(
                    {"error": "Permission denied - cannot manage sharing"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            action = request.data.get('action')
            
            if action == 'make_public':
                schema.sharing_level = 'public'
                schema.save()
                
                SchemaHistory.objects.create(
                    schema=schema,
                    action='shared',
                    user=request.user,
                    description="Made schema public"
                )
                
            elif action == 'make_org_shared':
                schema.sharing_level = 'org'
                schema.save()
                
                SchemaHistory.objects.create(
                    schema=schema,
                    action='shared',
                    user=request.user,
                    description="Shared schema with organization"
                )
                
            elif action == 'make_private':
                schema.sharing_level = 'private'
                schema.save()
                
                SchemaHistory.objects.create(
                    schema=schema,
                    action='unshared',
                    user=request.user,
                    description="Made schema private"
                )
                
            elif action == 'grant_permission':
                # Grant specific user permission
                target_user_id = request.data.get('user_id')
                permission_level = request.data.get('permission', 'view')
                
                try:
                    target_user = User.objects.get(id=target_user_id)
                    
                    permission, created = SchemaPermission.objects.get_or_create(
                        schema=schema,
                        user=target_user,
                        defaults={
                            'permission': permission_level,
                            'granted_by': request.user
                        }
                    )
                    
                    if not created:
                        permission.permission = permission_level
                        permission.granted_by = request.user
                        permission.save()
                    
                    SchemaHistory.objects.create(
                        schema=schema,
                        action='shared',
                        user=request.user,
                        description=f"Granted {permission_level} permission to {target_user.username}"
                    )
                    
                except User.DoesNotExist:
                    return Response(
                        {"error": "Target user not found"},
                        status=status.HTTP_404_NOT_FOUND
                    )
            
            # Return updated schema
            serializer = UserTableSchemaSerializer(schema, context={'request': request})
            return Response(serializer.data)
            
        except UserTableSchema.DoesNotExist:
            return Response(
                {"error": "Schema not found"},
                status=status.HTTP_404_NOT_FOUND
            )


class SchemaValidationView(CombinedOrgMixin, APIView):
    """Schema validation endpoint"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsReadOnlyOrAbove]

    def post(self, request, schema_id):
        """Validate schema and return detailed report"""
        try:
            schema = UserTableSchema.objects.get(id=schema_id)
            
            # Run validation
            is_valid = schema.validate_schema()
            
            return Response({
                "is_valid": is_valid,
                "errors": schema.validation_errors,
                "warnings": self.get_warnings(schema),
                "suggestions": self.get_suggestions(schema)
            })
            
        except UserTableSchema.DoesNotExist:
            return Response(
                {"error": "Schema not found"},
                status=status.HTTP_404_NOT_FOUND
            )
    
    def get_warnings(self, schema):
        """Get validation warnings"""
        warnings = []
        
        # Check for potential data loss scenarios
        if schema.sharing_level == 'public':
            warnings.append("Public schemas are visible to all users")
        
        # Check column types
        text_columns = schema.schema_columns.filter(data_type='text', max_length__isnull=True)
        if text_columns.exists():
            warnings.append(f"{text_columns.count()} text columns have no max_length set")
        
        return warnings
    
    def get_suggestions(self, schema):
        """Get optimization suggestions"""
        suggestions = []
        
        # Suggest adding indexes
        if not schema.schema_columns.filter(is_primary_key=True).exists():
            suggestions.append("Add a primary key column for better performance")
        
        # Suggest adding descriptions
        undocumented = schema.schema_columns.filter(description__isnull=True)
        if undocumented.exists():
            suggestions.append(f"Add descriptions to {undocumented.count()} columns for better documentation")
        
        return suggestions


class SharedSchemasView(CombinedOrgMixin, APIView):
    """Handle organization-wide shared schemas."""
    permission_classes = [CanAccessSharedSchemas]

    def get(self, request):
        """Get all organization-wide shared schemas."""
        shared_schemas = UserTableSchema.objects.filter(
            org=request.user.org,
            sharing_level='organization',
            is_shared=True
        )
        
        # Normalize columns for each schema
        for schema in shared_schemas:
            schema.columns = normalize_columns(schema.columns)
        
        serializer = UserTableSchemaSerializer(shared_schemas, many=True, context={'request': request})
        return Response(serializer.data)


class SchemaShareView(CombinedOrgMixin, APIView):
    """Handle sharing and unsharing of schemas."""
    permission_classes = [CanShareSchemas]

    def post(self, request, table_name):
        """Share a schema organization-wide."""
        try:
            schema = UserTableSchema.objects.get(
                user=request.user,
                org=request.user.org,
                table_name=table_name
            )
        except UserTableSchema.DoesNotExist:
            return Response({"error": "Schema not found."}, status=404)
        
        if not schema.can_user_share(request.user):
            return Response({
                "error": "You don't have permission to share this schema."
            }, status=403)
        
        if schema.is_shared:
            return Response({"error": "Schema is already shared."}, status=400)
        
        schema.share_organization_wide(request.user)
        
        # Log the activity
        UserActivity.objects.create(
            user=request.user,
            verb=f"shared schema '{table_name}' organization-wide",
            target=f"schema:{table_name}",
            meta={"table_name": table_name, "org_id": request.user.org.id}
        )
        
        serializer = UserTableSchemaSerializer(schema, context={'request': request})
        return Response(serializer.data)

    def delete(self, request, table_name):
        """Unshare a schema (make it personal)."""
        try:
            schema = UserTableSchema.objects.get(
                user=request.user,
                org=request.user.org,
                table_name=table_name
            )
        except UserTableSchema.DoesNotExist:
            return Response({"error": "Schema not found."}, status=404)
        
        if not schema.can_user_share(request.user):
            return Response({
                "error": "You don't have permission to unshare this schema."
            }, status=403)
        
        if not schema.is_shared:
            return Response({"error": "Schema is not currently shared."}, status=400)
        
        schema.make_personal()
        
        # Log the activity
        UserActivity.objects.create(
            user=request.user,
            verb=f"made schema '{table_name}' personal",
            target=f"schema:{table_name}",
            meta={"table_name": table_name, "org_id": request.user.org.id}
        )
        
        serializer = UserTableSchemaSerializer(schema, context={'request': request})
        return Response(serializer.data)

